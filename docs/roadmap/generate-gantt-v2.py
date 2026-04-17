#!/usr/bin/env python3
"""
Generate the Nova Gantt Chart v2 — styled Excel with integration deep-dive.
Run: python3 docs/roadmap/generate-gantt-v2.py
Output: docs/roadmap/Nova_Gantt_Integrations_v2.xlsx
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ══════════════════════════════════════════════════════════════════════
# COLORS
# ══════════════════════════════════════════════════════════════════════
CHARCOAL = "32373C"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
BORDER_GRAY = "D9D9D9"
HEADER_BG = "32373C"

TRACK_COLORS = {
    1: ("E8F0FE", "D2E3FC"),  # blue — Pipeline & Quality
    2: ("F3E8FD", "E1D5F0"),  # purple — Microsoft
    3: ("FFF8E1", "FFEEBA"),  # yellow — AI/Model Providers
    4: ("E6F4EA", "CEEAD6"),  # green — External Platforms
    5: ("FFF3E0", "FFE0B2"),  # orange — Analytics
    6: ("FCE4EC", "F8BBD0"),  # pink — Team Enablement
    7: ("E0F7FA", "B2EBF2"),  # cyan — China Eng
    8: ("F3E5F5", "CE93D8"),  # deep purple — Leadership
}

STATUS_FILLS = {
    "Done": "C8E6C9",
    "Active": "BBDEFB",
    "In Progress": "BBDEFB",
    "Planned": "F5F5F5",
    "Pending IT": "FFF9C4",
    "Requested": "FFF9C4",
    "Not Started": "F5F5F5",
    "Blocked": "FFCDD2",
    "Credits needed": "FFF9C4",
    "Partial": "BBDEFB",
    "Active (dev keys)": "BBDEFB",
    "Active (test)": "BBDEFB",
    "Pending": "FFF9C4",
    "Planned (Phase 2)": "F5F5F5",
}

# ══════════════════════════════════════════════════════════════════════
# FONTS & STYLES
# ══════════════════════════════════════════════════════════════════════
title_font = Font(name="Segoe UI", size=18, bold=True, color=CHARCOAL)
subtitle_font = Font(name="Segoe UI", size=11, italic=True, color="737373")
date_font = Font(name="Segoe UI", size=10, color="737373")
header_font = Font(name="Segoe UI", size=10, bold=True, color=WHITE)
track_font = Font(name="Segoe UI", size=10, bold=True, color=CHARCOAL)
wbs_font = Font(name="Segoe UI", size=9, bold=True, color="6B21A8")
task_font = Font(name="Segoe UI", size=9, color="333333")
task_bold = Font(name="Segoe UI", size=9, bold=True, color="333333")
note_font = Font(name="Segoe UI", size=8, color="777777")
status_font = Font(name="Segoe UI", size=8, bold=True, color="333333")

thin_border = Border(
    left=Side(style="thin", color=BORDER_GRAY),
    right=Side(style="thin", color=BORDER_GRAY),
    top=Side(style="thin", color=BORDER_GRAY),
    bottom=Side(style="thin", color=BORDER_GRAY),
)

wrap_top = Alignment(wrap_text=True, vertical="top")
center_wrap = Alignment(wrap_text=True, vertical="center", horizontal="center")
center_top = Alignment(wrap_text=True, vertical="top", horizontal="center")


def style_cell(cell, font=task_font, fill=None, alignment=None, border=thin_border):
    cell.font = font
    if fill:
        cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
    if alignment:
        cell.alignment = alignment
    else:
        cell.alignment = wrap_top
    if border:
        cell.border = border


# ══════════════════════════════════════════════════════════════════════
# SHEET 1: GANTT CHART
# ══════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Gantt Chart"

# Column widths
col_config = {
    "A": 6,    # WBS
    "B": 42,   # Task Name
    "C": 16,   # Owner
    "D": 12,   # Status
    "E": 30,   # Integration
    "F": 8,    # Days
    "G": 12,   # Start
    "H": 12,   # End
    "I": 18,   # W1 Apr 21-25
    "J": 18,   # W2 Apr 28-May 2
    "K": 18,   # W3 May 5-9
    "L": 18,   # W4 May 12-16
    "M": 18,   # W5 May 19-23
    "N": 18,   # W6 May 26-30
}
for col, width in col_config.items():
    ws.column_dimensions[col].width = width

# ── Title block ────────────────────────────────────────────────────
ws.merge_cells("A1:N1")
c = ws["A1"]
c.value = "Nova — Gantt Chart & Integration Map"
c.font = title_font
c.alignment = Alignment(vertical="center")

ws.merge_cells("A2:N2")
c = ws["A2"]
c.value = "8 Tracks  •  134 Tasks  •  25 Integrations  •  Business Days Only (No Weekends)"
c.font = subtitle_font

ws.merge_cells("A3:N3")
c = ws["A3"]
c.value = "Prepared: April 17, 2026  •  Start: April 21, 2026 (Monday)  •  Target End: May 28, 2026"
c.font = date_font

ws.row_dimensions[1].height = 32
ws.row_dimensions[2].height = 20
ws.row_dimensions[3].height = 18
ws.row_dimensions[4].height = 6  # spacer

# ── Headers (row 5) ───────────────────────────────────────────────
headers = [
    "WBS", "Task Name", "Owner", "Status", "Integration / Dependency",
    "Days", "Start", "End",
    "Week 1\nApr 21–25", "Week 2\nApr 28–May 2", "Week 3\nMay 5–9",
    "Week 4\nMay 12–16", "Week 5\nMay 19–23", "Week 6\nMay 26–30",
]

header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
for i, h in enumerate(headers, 1):
    c = ws.cell(row=5, column=i, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center_wrap
    c.border = thin_border
ws.row_dimensions[5].height = 36

# ── Task data ─────────────────────────────────────────────────────
# (track, wbs, task, owner, status, integration, days, start, end, w1-w6)

# Week date ranges for Gantt bar mapping
from datetime import date, timedelta

WEEKS = [
    (date(2026, 4, 21), date(2026, 4, 25)),
    (date(2026, 4, 28), date(2026, 5, 2)),
    (date(2026, 5, 5), date(2026, 5, 9)),
    (date(2026, 5, 12), date(2026, 5, 16)),
    (date(2026, 5, 19), date(2026, 5, 23)),
    (date(2026, 5, 26), date(2026, 5, 30)),
]


def week_bars(start_str, end_str):
    """Return list of 6 strings — filled marker for weeks that overlap."""
    try:
        s = date.fromisoformat(start_str)
        e = date.fromisoformat(end_str)
    except (ValueError, TypeError):
        return [""] * 6
    result = []
    for ws_date, we_date in WEEKS:
        if s <= we_date and e >= ws_date:
            result.append("■")
        else:
            result.append("")
    return result


def biz_end(start_str, days):
    """Add business days to a date string."""
    if days <= 1:
        return start_str
    d = date.fromisoformat(start_str)
    added = 0
    while added < days - 1:
        d += timedelta(days=1)
        if d.weekday() < 5:
            added += 1
    return d.isoformat()


tasks = [
    # ─── Track 1: Pipeline & Quality ───────────────────────────────
    (1, "1", "Pipeline & Quality", "", "", "", 20, "2026-04-21", "2026-05-18"),
    (1, "1.1", "Stage 4 — HTML Template Tuning", "Steven", "In Progress", "", 5, "2026-04-21", "2026-04-25"),
    (1, "1.1.1", "Audit existing templates + grading results", "Steven", "In Progress", "", 1, "2026-04-21", "2026-04-21"),
    (1, "1.1.2", "New feed format templates (3 variants)", "Steven", "Planned", "GLM-5 via OpenRouter", 1, "2026-04-22", "2026-04-22"),
    (1, "1.1.3", "New story/carousel templates (3 variants)", "Steven", "Planned", "GLM-5 via OpenRouter", 1, "2026-04-23", "2026-04-23"),
    (1, "1.1.4", "Typography + spacing hierarchy system", "Steven", "Planned", "", 1, "2026-04-24", "2026-04-24"),
    (1, "1.1.5", "Test all against 8-cat grading matrix", "Steven", "Planned", "Playwright (Chromium)", 1, "2026-04-25", "2026-04-25"),

    (1, "1.2", "Stage 6 — Landing Page Template Design", "Steven", "In Progress", "", 5, "2026-04-21", "2026-04-25"),
    (1, "1.2.1", "Hero section variants (3 concepts)", "Steven", "In Progress", "", 1, "2026-04-21", "2026-04-21"),
    (1, "1.2.2", "Body sections (testimonials, FAQ, benefits)", "Steven", "Planned", "", 1, "2026-04-22", "2026-04-22"),
    (1, "1.2.3", "Mobile-responsive + CTA placement", "Steven", "Planned", "", 1, "2026-04-23", "2026-04-23"),
    (1, "1.2.4", "Per-persona visual adaptation logic", "Steven", "Planned", "Vercel Blob", 1, "2026-04-24", "2026-04-24"),
    (1, "1.2.5", "Template selection logic in Stage 6", "Steven", "Planned", "Jinja2", 1, "2026-04-25", "2026-04-25"),

    (1, "1.3", "Stage 3 — Email + Job Posting Copy", "Steven", "Planned", "", 3, "2026-04-28", "2026-04-30"),
    (1, "1.3.1", "Add email sequence output to Stage 3", "Steven", "Planned", "Kimi K2.5 via OpenRouter", 1, "2026-04-28", "2026-04-28"),
    (1, "1.3.2", "Add job posting copy output", "Steven", "Planned", "Kimi K2.5 via OpenRouter", 1, "2026-04-29", "2026-04-29"),
    (1, "1.3.3", "Surface email/JD copy in marketing view", "Steven", "Planned", "", 1, "2026-04-30", "2026-04-30"),

    (1, "1.4", "Stage 2 — Persona Fix + Regeneration", "Steven", "Planned", "", 2, "2026-04-28", "2026-04-29"),
    (1, "1.4.1", "Regen Stage 2 for existing campaigns", "Steven", "Planned", "Seedream 4.5 via NIM", 1, "2026-04-28", "2026-04-28"),
    (1, "1.4.2", "Verify 3 actors/persona + grouping", "Steven", "Planned", "", 1, "2026-04-29", "2026-04-29"),

    (1, "1.5", "Organic Content (Posts, Flyers, Posters)", "Steven", "Planned", "", 4, "2026-05-01", "2026-05-06"),
    (1, "1.5.1", "Organic copy prompts (LI, FB, X)", "Steven", "Planned", "Kimi K2.5 via OpenRouter", 1, "2026-05-01", "2026-05-01"),
    (1, "1.5.2", "Print format dimensions (A4, A3, card)", "Steven", "Planned", "", 1, "2026-05-04", "2026-05-04"),
    (1, "1.5.3", "Composition templates for print", "Steven", "Planned", "GLM-5 + Playwright", 1, "2026-05-05", "2026-05-05"),
    (1, "1.5.4", "Gallery integration + designer flow", "Steven", "Planned", "", 1, "2026-05-06", "2026-05-06"),

    (1, "1.6", "Stage 5 — Video Pipeline", "Steven", "Planned", "", 3, "2026-05-07", "2026-05-11"),
    (1, "1.6.1", "Kling credits + test on real campaign", "Steven", "Planned", "Kling 3.0 API", 1, "2026-05-07", "2026-05-07"),
    (1, "1.6.2", "UGC script gen + storyboard", "Steven", "Planned", "Kimi K2.5 via OpenRouter", 1, "2026-05-08", "2026-05-08"),
    (1, "1.6.3", "TTS + lip sync + FFmpeg", "Steven", "Planned", "Coqui TTS + Wav2Lip", 1, "2026-05-11", "2026-05-11"),

    (1, "1.7", "End-to-End Campaign Testing", "Steven + Jenn", "Planned", "", 5, "2026-05-12", "2026-05-18"),
    (1, "1.7.1", "Campaign 1 — full pipeline", "Steven + Jenn", "Planned", "All integrations", 1, "2026-05-12", "2026-05-12"),
    (1, "1.7.2", "Campaign 2 — different region/lang", "Steven + Jenn", "Planned", "", 1, "2026-05-13", "2026-05-13"),
    (1, "1.7.3", "Campaign 3 — onsite task type", "Steven + Jenn", "Planned", "", 1, "2026-05-14", "2026-05-14"),
    (1, "1.7.4", "Bug fixes from testing", "Steven", "Planned", "", 1, "2026-05-15", "2026-05-15"),
    (1, "1.7.5", "Quality sign-off", "Steven + Jenn", "Planned", "", 1, "2026-05-18", "2026-05-18"),

    # ─── Track 2: Microsoft Integrations ───────────────────────────
    (2, "2", "Microsoft Integrations", "", "", "", 15, "2026-04-21", "2026-05-04"),
    (2, "2.1", "Azure AD SSO (SAML)", "IT + Steven", "Not Started", "Azure AD + Clerk SAML", 5, "2026-04-21", "2026-04-25"),
    (2, "2.1.1", "IT: Create Azure AD enterprise app", "IT Admin", "Not Started", "Azure Active Directory", 2, "2026-04-21", "2026-04-22"),
    (2, "2.1.2", "IT: Provide SAML metadata URL", "IT Admin", "Not Started", "Azure AD SAML", 1, "2026-04-23", "2026-04-23"),
    (2, "2.1.3", "Configure Clerk SAML connection", "Steven", "Not Started", "Clerk SSO", 1, "2026-04-24", "2026-04-24"),
    (2, "2.1.4", "Test SSO login + role provisioning", "Steven", "Not Started", "Clerk + Azure AD", 1, "2026-04-25", "2026-04-25"),

    (2, "2.2", "Microsoft Teams Notifications", "IT + Steven", "Partial", "Teams Webhooks", 3, "2026-04-21", "2026-04-23"),
    (2, "2.2.1", "IT: Create prod webhook URL", "IT Admin", "Not Started", "Teams Admin Center", 1, "2026-04-21", "2026-04-21"),
    (2, "2.2.2", "Swap test → prod webhook URL", "Steven", "Not Started", "Vercel env vars", 1, "2026-04-22", "2026-04-22"),
    (2, "2.2.3", "Test adaptive card delivery E2E", "Steven", "Not Started", "Teams Webhooks", 1, "2026-04-23", "2026-04-23"),

    (2, "2.3", "Outlook Email Notifications", "IT + Steven", "Not Started", "Microsoft Graph API", 3, "2026-04-24", "2026-04-28"),
    (2, "2.3.1", "IT: Register Graph API app (Mail.Send)", "IT Admin", "Not Started", "Azure App Registration", 1, "2026-04-24", "2026-04-24"),
    (2, "2.3.2", "Configure Graph API creds in Nova", "Steven", "Not Started", "Microsoft Graph SDK", 1, "2026-04-25", "2026-04-25"),
    (2, "2.3.3", "Test email from @centific.com", "Steven", "Not Started", "Graph API + Outlook", 1, "2026-04-28", "2026-04-28"),

    (2, "2.4", "SharePoint Campaign Folders", "IT + Steven", "Not Started", "Microsoft Graph API", 4, "2026-04-29", "2026-05-02"),
    (2, "2.4.1", "IT: Grant Sites.ReadWrite.All", "IT Admin", "Not Started", "Azure App Permissions", 1, "2026-04-29", "2026-04-29"),
    (2, "2.4.2", "IT: Confirm site + folder structure", "IT + Jenn", "Not Started", "SharePoint Online", 1, "2026-04-30", "2026-04-30"),
    (2, "2.4.3", "Build auto-folder on approval", "Steven", "Not Started", "Graph API (DriveItems)", 1, "2026-05-01", "2026-05-01"),
    (2, "2.4.4", "Auto-upload approved assets", "Steven", "Not Started", "Graph API + Blob", 1, "2026-05-02", "2026-05-02"),

    # ─── Track 3: AI/Model Providers ───────────────────────────────
    (3, "3", "AI & Model Providers", "", "", "", 10, "2026-04-21", "2026-05-11"),
    (3, "3.1", "OpenRouter (LLM Gateway)", "Steven", "Active", "OpenRouter API", 2, "2026-04-21", "2026-04-22"),
    (3, "3.1.1", "Verify Kimi K2.5 quota (Stages 1, 3)", "Steven", "Active", "OpenRouter", 1, "2026-04-21", "2026-04-21"),
    (3, "3.1.2", "Verify GLM-5 quota (Stage 4)", "Steven", "Active", "OpenRouter", 1, "2026-04-22", "2026-04-22"),

    (3, "3.2", "NVIDIA NIM (Image Gen)", "Steven", "Active", "NVIDIA NIM API", 2, "2026-04-21", "2026-04-22"),
    (3, "3.2.1", "Verify Seedream 4.5 credits", "Steven", "Active", "NIM — Stage 2", 1, "2026-04-21", "2026-04-21"),
    (3, "3.2.2", "Test 15-key rotation", "Steven", "Active", "NIM key rotation", 1, "2026-04-22", "2026-04-22"),

    (3, "3.3", "Kling 3.0 (Video Gen)", "Steven", "Planned", "Kling API", 2, "2026-05-07", "2026-05-08"),
    (3, "3.3.1", "Load credits + verify API", "Steven", "Planned", "Kling — Stage 5", 1, "2026-05-07", "2026-05-07"),
    (3, "3.3.2", "Test multi-shot video gen", "Steven", "Planned", "Kling + FFmpeg", 1, "2026-05-08", "2026-05-08"),

    (3, "3.4", "MLX Local Inference", "Steven", "Active", "MLX + Metal GPU", 1, "2026-04-23", "2026-04-23"),
    (3, "3.5", "Gemma 4 (VQA)", "Steven", "Active", "Gemma via NIM/OR", 1, "2026-04-24", "2026-04-24"),
    (3, "3.6", "Coqui TTS + Wav2Lip", "Steven", "Planned", "Local torch + opencv", 1, "2026-05-11", "2026-05-11"),

    # ─── Track 4: External Platforms ───────────────────────────────
    (4, "4", "External Platforms", "", "", "", 8, "2026-04-21", "2026-04-28"),
    (4, "4.1", "WordPress (Job CPT + ACF)", "Steven", "Done", "WP REST API + ACF", 0, "2026-04-21", "2026-04-21"),
    (4, "4.1.1", "Job CPT auto-publish", "Steven", "Done", "WP REST API v2", 0, "2026-04-21", "2026-04-21"),
    (4, "4.1.2", "ACF apply_job repeater", "Steven", "Done", "ACF REST API", 0, "2026-04-21", "2026-04-21"),
    (4, "4.1.3", "Yoast SEO meta via excerpt", "Steven", "Done", "Yoast + WP REST", 0, "2026-04-21", "2026-04-21"),
    (4, "4.1.4", "Taxonomy tagging", "Steven", "Done", "WP Taxonomies", 0, "2026-04-21", "2026-04-21"),

    (4, "4.2", "Vercel (Hosting + Blob)", "Steven", "Active", "Vercel Platform", 2, "2026-04-21", "2026-04-22"),
    (4, "4.2.1", "Production deploy + CI", "Steven", "Done", "Vercel CLI + GH Actions", 0, "2026-04-21", "2026-04-21"),
    (4, "4.2.2", "go.oneforma.com domain", "IT + Steven", "Pending IT", "DNS CNAME + Vercel", 2, "2026-04-21", "2026-04-22"),

    (4, "4.3", "Neon Postgres", "Steven", "Active", "Neon Serverless", 1, "2026-04-21", "2026-04-21"),
    (4, "4.3.1", "DB operational", "Steven", "Done", "@neondatabase + asyncpg", 0, "2026-04-21", "2026-04-21"),
    (4, "4.3.2", "Rotate DB password", "Steven", "Pending", "Neon Console", 1, "2026-04-21", "2026-04-21"),

    (4, "4.4", "Clerk (Auth)", "Steven", "Active", "Clerk SDK", 1, "2026-04-25", "2026-04-25"),
    (4, "4.4.1", "Swap to prod keys (after SSO)", "Steven", "Blocked", "Clerk + Azure AD", 1, "2026-04-25", "2026-04-25"),

    (4, "4.5", "Figma (Design Handoff)", "Steven + Miguel", "Active", "Figma API", 1, "2026-04-28", "2026-04-28"),
    (4, "4.5.1", "Test push-to-Figma on live campaign", "Miguel", "Planned", "Figma PAT", 1, "2026-04-28", "2026-04-28"),

    (4, "4.6", "Playwright (Rendering)", "Steven", "Done", "Chromium headless", 0, "2026-04-21", "2026-04-21"),

    # ─── Track 5: Analytics ────────────────────────────────────────
    (5, "5", "Analytics & Ad Platforms", "", "", "", 10, "2026-05-18", "2026-05-28"),
    (5, "5.1", "Google Analytics 4", "Steven", "Planned", "GA4 Data API", 3, "2026-05-18", "2026-05-20"),
    (5, "5.1.1", "Get GA4 property access", "Steven + Jenn", "Planned", "Google Cloud Console", 1, "2026-05-18", "2026-05-18"),
    (5, "5.1.2", "Connect GA4 API to Visualize", "Steven", "Planned", "GA4 API + OAuth", 1, "2026-05-19", "2026-05-19"),
    (5, "5.1.3", "UTM attribution mapping", "Steven", "Planned", "GA4 + tracked_links", 1, "2026-05-20", "2026-05-20"),

    (5, "5.2", "Meta Ads (FB/IG)", "Steven", "Planned", "Meta Marketing API", 2, "2026-05-21", "2026-05-22"),
    (5, "5.2.1", "Get Meta Business Manager access", "Steven + Jenn", "Planned", "Meta Business Suite", 1, "2026-05-21", "2026-05-21"),
    (5, "5.2.2", "Connect Meta API to Visualize", "Steven", "Planned", "Meta Marketing API v19", 1, "2026-05-22", "2026-05-22"),

    (5, "5.3", "LinkedIn Campaign Manager", "Steven", "Planned", "LinkedIn Marketing API", 2, "2026-05-25", "2026-05-26"),
    (5, "5.3.1", "Get LinkedIn ad account access", "Steven + Jenn", "Planned", "LinkedIn Campaign Mgr", 1, "2026-05-25", "2026-05-25"),
    (5, "5.3.2", "Connect LinkedIn API to Visualize", "Steven", "Planned", "LinkedIn Marketing API", 1, "2026-05-26", "2026-05-26"),

    (5, "5.4", "Google Ads", "Steven", "Planned", "Google Ads API", 2, "2026-05-27", "2026-05-28"),
    (5, "5.4.1", "Get Google Ads account access", "Steven + Jenn", "Planned", "Google Ads Console", 1, "2026-05-27", "2026-05-27"),
    (5, "5.4.2", "Connect Google Ads API to Visualize", "Steven", "Planned", "Google Ads API v16", 1, "2026-05-28", "2026-05-28"),

    # ─── Track 6: Team Enablement ──────────────────────────────────
    (6, "6", "Team Enablement", "", "", "", 15, "2026-04-28", "2026-05-14"),
    (6, "6.1", "Designer Onboarding (Miguel)", "Miguel", "Planned", "Figma + Designer Portal", 5, "2026-04-28", "2026-05-02"),
    (6, "6.1.1", "Portal walkthrough + feedback", "Steven + Miguel", "Planned", "", 1, "2026-04-28", "2026-04-28"),
    (6, "6.1.2", "Daily use on live campaigns (1 wk)", "Miguel", "Planned", "", 4, "2026-04-29", "2026-05-02"),

    (6, "6.2", "Recruiter Pilot (1-2 volunteers)", "Jenn", "Planned", "", 5, "2026-05-05", "2026-05-09"),
    (6, "6.2.1", "Identify pilot recruiters", "Jenn", "Planned", "", 1, "2026-05-05", "2026-05-05"),
    (6, "6.2.2", "Supervised intake walkthrough", "Steven", "Planned", "", 1, "2026-05-06", "2026-05-06"),
    (6, "6.2.3", "Pilots run real campaigns", "Recruiters", "Planned", "", 3, "2026-05-07", "2026-05-09"),

    (6, "6.3", "Full Team Training + Agency", "Steven", "Planned", "", 3, "2026-05-12", "2026-05-14"),
    (6, "6.3.1", "Record Loom walkthroughs (4 portals)", "Steven", "Planned", "", 1, "2026-05-12", "2026-05-12"),
    (6, "6.3.2", "Live training — full recruiter team", "Steven", "Planned", "", 1, "2026-05-13", "2026-05-13"),
    (6, "6.3.3", "Agency test handoff via magic link", "Steven + Jenn", "Planned", "", 1, "2026-05-14", "2026-05-14"),

    # ─── Track 7: China Engineering ────────────────────────────────
    (7, "7", "China Engineering Deployment", "", "", "", 8, "2026-04-21", "2026-04-28"),
    (7, "7.1", "Codebase Review + Architecture", "Michael", "In Progress", "GitHub (OneTake)", 3, "2026-04-21", "2026-04-23"),
    (7, "7.1.1", "Michael reviews codebase + README", "Michael", "In Progress", "", 2, "2026-04-21", "2026-04-22"),
    (7, "7.1.2", "Architecture decision", "Michael + Steven", "Planned", "", 1, "2026-04-23", "2026-04-23"),
    (7, "7.2", "Azure VM Deployment", "Michael", "Planned", "Docker + Azure", 3, "2026-04-24", "2026-04-28"),
    (7, "7.2.1", "Provision Azure VM + Docker", "Michael", "Planned", "Azure VM + Docker", 1, "2026-04-24", "2026-04-24"),
    (7, "7.2.2", "Deploy worker container", "Michael", "Planned", "Docker Compose", 1, "2026-04-25", "2026-04-25"),
    (7, "7.2.3", "E2E verification on Azure", "Michael + Steven", "Planned", "All integrations", 1, "2026-04-28", "2026-04-28"),

    # ─── Track 8: Leadership Pitch ─────────────────────────────────
    (8, "8", "Leadership Pitch", "", "", "", 8, "2026-05-18", "2026-05-28"),
    (8, "8.1", "Prepare Pitch Materials", "Steven", "Planned", "", 5, "2026-05-18", "2026-05-22"),
    (8, "8.1.1", "Compile campaign results + metrics", "Steven", "Planned", "Nova DB + GA4", 2, "2026-05-18", "2026-05-19"),
    (8, "8.1.2", "Build pitch deck", "Steven", "Planned", "", 2, "2026-05-20", "2026-05-21"),
    (8, "8.1.3", "Rehearse live demo flow", "Steven", "Planned", "", 1, "2026-05-22", "2026-05-22"),
    (8, "8.2", "Leadership Presentations", "Steven + Jenn", "Planned", "", 3, "2026-05-25", "2026-05-28"),
    (8, "8.2.1", "Informal demo to PM team", "Steven + Jenn", "Planned", "", 1, "2026-05-25", "2026-05-25"),
    (8, "8.2.2", "SVP Engineering presentation", "Steven", "Planned", "VYRA Visualize demo", 1, "2026-05-27", "2026-05-27"),
    (8, "8.2.3", "Follow-up + next steps", "Steven + Jenn", "Planned", "", 1, "2026-05-28", "2026-05-28"),
]

# ── Write task rows ───────────────────────────────────────────────
row_num = 6
current_track = None
group_starts = {}  # track_num -> (first_row, last_row) for grouping

for (track_n, wbs, task, owner, status, integration, days, start, end) in tasks:
    level = wbs.count(".")
    track_bg, track_ms = TRACK_COLORS[track_n]
    bars = week_bars(start, end)
    status_fill = STATUS_FILLS.get(status, LIGHT_GRAY)

    # Track grouping
    if track_n not in group_starts:
        group_starts[track_n] = {"start": row_num, "end": row_num}
    else:
        group_starts[track_n]["end"] = row_num

    # WBS
    c = ws.cell(row=row_num, column=1, value=wbs)
    style_cell(c, font=wbs_font, fill=track_bg, alignment=center_top)

    # Task Name (indented by level)
    indent = "    " * level
    c = ws.cell(row=row_num, column=2, value=f"{indent}{task}")
    font = track_font if level == 0 else (task_bold if level == 1 else task_font)
    style_cell(c, font=font, fill=track_bg)

    # Owner
    c = ws.cell(row=row_num, column=3, value=owner)
    style_cell(c, font=task_font, fill=track_bg, alignment=center_top)

    # Status
    c = ws.cell(row=row_num, column=4, value=status)
    style_cell(c, font=status_font, fill=status_fill, alignment=center_top)

    # Integration
    c = ws.cell(row=row_num, column=5, value=integration)
    style_cell(c, font=note_font, fill=track_bg)

    # Days
    c = ws.cell(row=row_num, column=6, value=days if days > 0 else "")
    style_cell(c, font=task_font, fill=track_bg, alignment=center_top)

    # Start
    c = ws.cell(row=row_num, column=7, value=start)
    style_cell(c, font=note_font, fill=track_bg, alignment=center_top)

    # End
    c = ws.cell(row=row_num, column=8, value=end)
    style_cell(c, font=note_font, fill=track_bg, alignment=center_top)

    # Week bars (cols I-N)
    for wi, bar in enumerate(bars):
        c = ws.cell(row=row_num, column=9 + wi, value=bar)
        fill = track_ms if bar else LIGHT_GRAY
        bar_color = TRACK_COLORS[track_n][1]
        style_cell(c, font=Font(name="Segoe UI", size=12, bold=True, color=bar_color),
                   fill=fill, alignment=center_wrap)

    # Row height
    ws.row_dimensions[row_num].height = 22 if level >= 2 else (26 if level == 1 else 28)
    row_num += 1

# ── Row grouping (outline) for collapsible sections ───────────────
for track_n, info in group_starts.items():
    first = info["start"]
    last = info["end"]
    # Group detail rows (skip the track header row)
    if last > first:
        for r in range(first + 1, last + 1):
            ws.row_dimensions[r].outline_level = 1

# Enable outline grouping
try:
    from openpyxl.worksheet.properties import OutlineProperties
    ws.sheet_properties.outlinePr = OutlineProperties(summaryBelow=False, summaryRight=False)
except (ImportError, AttributeError):
    pass  # Older openpyxl — grouping still works, just no summary config

# ── Freeze panes ──────────────────────────────────────────────────
ws.freeze_panes = "C6"

# ── Print setup ───────────────────────────────────────────────────
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0


# ══════════════════════════════════════════════════════════════════════
# SHEET 2: INTEGRATION MAP
# ══════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Integration Map")

int_col_widths = {"A": 16, "B": 24, "C": 45, "D": 38, "E": 16, "F": 40}
for col, w in int_col_widths.items():
    ws2.column_dimensions[col].width = w

# Title
ws2.merge_cells("A1:F1")
c = ws2["A1"]
c.value = "Nova — Integration Map (25 Services)"
c.font = title_font

ws2.merge_cells("A2:F2")
c = ws2["A2"]
c.value = "Every external service, API, and dependency — with credentials, status, and notes"
c.font = subtitle_font

ws2.row_dimensions[1].height = 32
ws2.row_dimensions[2].height = 20
ws2.row_dimensions[3].height = 6

# Headers
int_headers = ["Category", "Service", "Used For", "Credential / Config", "Status", "Notes"]
for i, h in enumerate(int_headers, 1):
    c = ws2.cell(row=4, column=i, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center_wrap
    c.border = thin_border
ws2.row_dimensions[4].height = 30

integrations = [
    ("AI / LLM", "OpenRouter", "Kimi K2.5 (Stages 1,3), GLM-5 (Stage 4)", "OPENROUTER_API_KEY", "Active", "Paid API — check quota monthly"),
    ("AI / Image", "NVIDIA NIM", "Seedream 4.5 image gen (Stage 2)", "NVIDIA_NIM_API_KEY (x15)", "Active", "Free tier keys, rotate monthly"),
    ("AI / VQA", "Gemma 4", "Visual quality assurance (Stages 2,4)", "Via OpenRouter or NIM", "Active", "Pass threshold: 0.85"),
    ("AI / Video", "Kling 3.0", "UGC video generation (Stage 5)", "KLING_ACCESS_KEY + SECRET", "Credits needed", "Paid API, scale requires credits"),
    ("AI / Local", "MLX (Apple Silicon)", "Local LLM/VLM fallback", "N/A (Metal GPU)", "Active", "Not available in Docker"),
    ("AI / Local", "Coqui TTS", "Text-to-speech (Stage 5)", "N/A (local)", "Planned", "XTTS-v2, 17 languages"),
    ("AI / Local", "Wav2Lip", "Lip sync for video (Stage 5)", "N/A (local torch)", "Planned", "Requires GPU"),
    ("CMS", "WordPress", "Job CPT, ACF repeater, Yoast SEO", "WP_USERNAME + WP_APP_PASSWORD", "Active", "REST API v2, ACF REST enabled"),
    ("Auth", "Clerk", "User auth, SSO, roles", "CLERK_SECRET_KEY + PUB_KEY", "Active (dev)", "Swap to prod after Azure AD"),
    ("Auth", "Azure AD (SAML)", "Enterprise SSO for Centific", "SAML metadata URL from IT", "Not Started", "Blocks internal rollout"),
    ("Database", "Neon Postgres", "Primary data store + job queue", "DATABASE_URL", "Active", "PASSWORD ROTATION NEEDED"),
    ("Storage", "Vercel Blob", "Image, HTML, file storage", "BLOB_READ_WRITE_TOKEN", "Active", "Public + private blobs"),
    ("Hosting", "Vercel", "Next.js hosting, serverless functions", "Vercel CLI auth", "Active", "nova-intake.vercel.app"),
    ("Hosting", "Azure VM", "Worker container (China team)", "Azure subscription", "Planned", "Docker multi-arch ready"),
    ("Notifications", "Microsoft Teams", "Adaptive card notifications", "TEAMS_WEBHOOK_URL", "Active (test)", "Need prod webhook from IT"),
    ("Notifications", "Outlook / Graph", "Email from @centific.com", "Graph API app registration", "Not Started", "Need IT to register app"),
    ("Design", "Figma", "Push creatives for designer", "Personal Access Token", "Active", "figma-api npm package"),
    ("Rendering", "Playwright", "HTML→PNG composition", "N/A (headless browser)", "Active", "Bundled in Docker image"),
    ("Rendering", "FFmpeg", "Video/audio muxing", "N/A (system binary)", "Active", "Bundled in Docker image"),
    ("DNS", "go.oneforma.com", "Branded short link domain", "CNAME → cname.vercel-dns.com", "Pending IT", "Cosmetic — works without it"),
    ("Analytics", "Google Analytics 4", "Campaign attribution", "GA4 property + service acct", "Planned (Phase 2)", "VYRA Visualize dependency"),
    ("Analytics", "Meta Marketing API", "FB/IG ad performance", "Meta Business Manager", "Planned (Phase 2)", "VYRA Visualize dependency"),
    ("Analytics", "LinkedIn Marketing", "LinkedIn ad performance", "LinkedIn Campaign Manager", "Planned (Phase 2)", "VYRA Visualize dependency"),
    ("Analytics", "Google Ads API", "Search/Display performance", "Google Ads account", "Planned (Phase 2)", "VYRA Visualize dependency"),
    ("File Storage", "SharePoint Online", "Auto-save approved assets", "Graph API Sites.ReadWrite", "Not Started", "P2 — nice to have"),
]

CAT_COLORS = {
    "AI / LLM": "FFF8E1", "AI / Image": "FFF8E1", "AI / VQA": "FFF8E1",
    "AI / Video": "FFF8E1", "AI / Local": "FFF8E1",
    "CMS": "E6F4EA", "Auth": "F3E8FD", "Database": "E8F0FE",
    "Storage": "E8F0FE", "Hosting": "E8F0FE",
    "Notifications": "FCE4EC", "Design": "F3E8FD",
    "Rendering": "E0F7FA", "DNS": "E0F7FA",
    "Analytics": "FFF3E0", "File Storage": "E6F4EA",
}

row = 5
for cat, svc, used, cred, status, notes in integrations:
    cat_fill = CAT_COLORS.get(cat, LIGHT_GRAY)
    status_fill = STATUS_FILLS.get(status, LIGHT_GRAY)

    c = ws2.cell(row=row, column=1, value=cat)
    style_cell(c, font=task_bold, fill=cat_fill, alignment=center_top)

    c = ws2.cell(row=row, column=2, value=svc)
    style_cell(c, font=task_bold, fill=cat_fill)

    c = ws2.cell(row=row, column=3, value=used)
    style_cell(c, font=task_font, fill=cat_fill)

    c = ws2.cell(row=row, column=4, value=cred)
    style_cell(c, font=Font(name="Consolas", size=8, color="555555"), fill=cat_fill)

    c = ws2.cell(row=row, column=5, value=status)
    style_cell(c, font=status_font, fill=status_fill, alignment=center_top)

    c = ws2.cell(row=row, column=6, value=notes)
    style_cell(c, font=note_font, fill=cat_fill)

    ws2.row_dimensions[row].height = 24
    row += 1

ws2.freeze_panes = "C5"

# ══════════════════════════════════════════════════════════════════════
# SAVE
# ══════════════════════════════════════════════════════════════════════
output = "/Users/stevenjunop/centric-intake/docs/roadmap/Nova_Gantt_Integrations_v2.xlsx"
wb.save(output)
print(f"Saved: {output}")
print(f"Tasks: {len(tasks)}")
print(f"Integrations: {len(integrations)}")
